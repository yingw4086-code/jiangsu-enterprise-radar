from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO, Protocol, runtime_checkable

from openpyxl import load_workbook

if TYPE_CHECKING:
    from app.company_registry import CompanyRegistryRecord


EXCEL_HEADER_MAPPING = {
    "企业名称": "company_name",
    "统一社会信用代码": "unified_social_credit_code",
    "法人": "legal_person",
    "注册资本": "registered_capital",
    "成立日期": "establish_date",
    "注册地址": "company_address",
    "经营范围": "business_scope",
    "企业状态": "company_status",
    "行业": "industry",
}
EXCEL_HEADER_ALIASES = {
    "company_name": ("企业名称",),
    "unified_social_credit_code": ("统一社会信用代码",),
    "legal_person": ("法人",),
    "registered_capital": ("注册资本",),
    "establish_date": ("成立日期",),
    "company_address": ("注册地址",),
    "business_scope": ("经营范围",),
    "company_status": ("企业状态",),
    "industry": ("行业分类", "行业"),
}
MAX_EXCEL_BYTES = 10 * 1024 * 1024
MAX_EXCEL_ROWS = 5000


class CompanyDataProviderError(RuntimeError):
    pass


class CompanyDataProviderNotConfigured(CompanyDataProviderError):
    pass


class CompanyExcelValidationError(CompanyDataProviderError):
    pass


@runtime_checkable
class CompanyDataProvider(Protocol):
    provider_name: str

    def lookup(self, company_name: str) -> "CompanyRegistryRecord | None": ...


class NationalEnterpriseCreditProvider:
    """Reserved adapter for the national public registry; performs no I/O yet."""

    provider_name = "national_enterprise_credit_publicity"

    def lookup(self, company_name: str) -> "CompanyRegistryRecord | None":
        raise CompanyDataProviderNotConfigured(
            "国家企业信用信息公示系统适配器尚未配置，当前不会发起网络请求"
        )


class ThirdPartyCompanyDataProvider:
    """Reserved adapter for a user-supplied third-party client."""

    provider_name = "third_party_company_registry"

    def lookup(self, company_name: str) -> "CompanyRegistryRecord | None":
        raise CompanyDataProviderNotConfigured(
            "第三方工商数据接口尚未配置，当前不会调用收费 API"
        )


@dataclass(frozen=True)
class ExcelCompanyRecord:
    row_number: int
    record: "CompanyRegistryRecord"


class ExcelCompanyDataProvider:
    provider_name = "user_excel_import"

    def __init__(self, source: bytes | bytearray | Path | BinaryIO):
        self._source = source
        self._records: list[ExcelCompanyRecord] | None = None

    def load_records(self) -> list[ExcelCompanyRecord]:
        if self._records is None:
            self._records = _read_excel_records(self._source)
        return list(self._records)

    def lookup(self, company_name: str) -> "CompanyRegistryRecord | None":
        from app.company_matcher import CompanyRegistryMatcher

        records = [item.record for item in self.load_records()]
        return CompanyRegistryMatcher(records).match(company_name).record


def _read_excel_records(
    source: bytes | bytearray | Path | BinaryIO,
) -> list[ExcelCompanyRecord]:
    from app.company_matcher import normalize_company_name
    from app.company_registry import CompanyRegistryRecord

    workbook_source = _workbook_source(source)
    try:
        workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    except Exception as exc:
        raise CompanyExcelValidationError(f"无法读取 Excel 文件：{exc}") from exc
    try:
        sheet = workbook.active
        header_values = [str(cell.value or "").strip() for cell in sheet[1]]
        missing_headers = [
            "（或）".join(aliases)
            for aliases in EXCEL_HEADER_ALIASES.values()
            if not any(header in header_values for header in aliases)
        ]
        if missing_headers:
            raise CompanyExcelValidationError(
                "Excel 缺少字段：" + "、".join(missing_headers)
            )
        indexes = {
            field_name: next(
                header_values.index(header)
                for header in aliases
                if header in header_values
            )
            for field_name, aliases in EXCEL_HEADER_ALIASES.items()
        }
        records: list[ExcelCompanyRecord] = []
        seen_names: set[str] = set()
        seen_credit_codes: set[str] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row_number > MAX_EXCEL_ROWS + 1:
                raise CompanyExcelValidationError(
                    f"Excel 数据超过 {MAX_EXCEL_ROWS} 行限制"
                )
            values = {
                field: row[index].value if index < len(row) else None
                for field, index in indexes.items()
            }
            if not any(_cell_text(value) for value in values.values()):
                continue
            company_name = _cell_text(values["company_name"])
            if not company_name:
                raise CompanyExcelValidationError(f"第 {row_number} 行企业名称为空")
            normalized_name = normalize_company_name(company_name)
            if normalized_name in seen_names:
                raise CompanyExcelValidationError(
                    f"第 {row_number} 行企业名称重复：{company_name}"
                )
            seen_names.add(normalized_name)
            credit_code = _cell_text(values["unified_social_credit_code"]).upper()
            if credit_code and credit_code in seen_credit_codes:
                raise CompanyExcelValidationError(
                    f"第 {row_number} 行统一社会信用代码重复：{credit_code}"
                )
            if credit_code:
                seen_credit_codes.add(credit_code)
            records.append(
                ExcelCompanyRecord(
                    row_number=row_number,
                    record=CompanyRegistryRecord(
                        company_name=company_name,
                        unified_social_credit_code=credit_code,
                        legal_person=_cell_text(values["legal_person"]),
                        registered_capital=_cell_text(values["registered_capital"]),
                        establish_date=_date_text(
                            values["establish_date"], row_number=row_number
                        ),
                        company_address=_cell_text(values["company_address"]),
                        business_scope=_cell_text(values["business_scope"]),
                        company_status=_cell_text(values["company_status"]),
                        industry=_cell_text(values["industry"]),
                        data_source="user_excel_import",
                        verified_at=date.today().isoformat(),
                    ),
                )
            )
        if not records:
            raise CompanyExcelValidationError("Excel 中没有可导入的企业记录")
        return records
    finally:
        workbook.close()


def _workbook_source(source: bytes | bytearray | Path | BinaryIO) -> Any:
    if isinstance(source, (bytes, bytearray)):
        if len(source) > MAX_EXCEL_BYTES:
            raise CompanyExcelValidationError("Excel 文件不能超过 10MB")
        return BytesIO(bytes(source))
    if isinstance(source, Path):
        if source.stat().st_size > MAX_EXCEL_BYTES:
            raise CompanyExcelValidationError("Excel 文件不能超过 10MB")
        return source
    raw = source.read()
    if not isinstance(raw, (bytes, bytearray)):
        raise CompanyExcelValidationError("Excel 上传内容必须为二进制文件")
    if len(raw) > MAX_EXCEL_BYTES:
        raise CompanyExcelValidationError("Excel 文件不能超过 10MB")
    return BytesIO(bytes(raw))


def _date_text(value: Any, *, row_number: int) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _cell_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    raise CompanyExcelValidationError(
        f"第 {row_number} 行成立日期格式错误，应为 YYYY-MM-DD"
    )


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
