from __future__ import annotations

import sqlite3
import tempfile
import unittest
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.company_matcher import normalize_company_name
from app.company_registry import (
    enrich_items_with_company_registry,
    enrich_registry_completeness,
    import_company_registry_excel,
    list_company_registry_records,
    summarize_registry_coverage,
)
from app.enterprise_profile_enhance import (
    build_enhanced_enterprise_profile,
    enrich_company_strength,
)
from app.finance_scoring import enrich_finance_opportunities, score_finance_opportunity
from app.marketing_report import build_marketing_report
from app.official_permit_data import load_official_permit_dataset
from app.permit_data_runtime import load_planning_permit_dataset


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PRODUCTION_DB = PROJECT_ROOT / "database" / "enterprise.db"
TEMPLATE_PATH = (
    PROJECT_ROOT
    / "outputs"
    / "phase3_10"
    / "company_registry_import_template.xlsx"
)
HEADERS = [
    "企业名称",
    "统一社会信用代码",
    "法人",
    "注册资本",
    "成立日期",
    "注册地址",
    "经营范围",
    "企业状态",
    "行业",
]


def _workbook_bytes(company_names: list[str]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for index, company_name in enumerate(company_names, start=1):
        sheet.append(
            [
                company_name,
                f"91320684TS{index:08d}",
                f"测试法人{index:02d}",
                "1亿元",
                datetime(2010, 1, index),
                f"海门区测试注册地址{index:02d}号",
                "工业设备研发、生产与销售",
                "存续",
                "制造业",
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _company_name(item: dict[str, object]) -> str:
    return str(
        item.get("company_name")
        or item.get("construction_unit")
        or item.get("owner_name")
        or ""
    ).strip()


def _copy_database(source: Path, destination: Path) -> None:
    source_connection = sqlite3.connect(source)
    destination_connection = sqlite3.connect(destination)
    try:
        source_connection.backup(destination_connection)
    finally:
        destination_connection.close()
        source_connection.close()


def _clear_registry_fixture(db_path: Path) -> None:
    connection = sqlite3.connect(db_path)
    try:
        with connection:
            connection.execute("DELETE FROM company_registry")
            log_table = connection.execute(
                """
                SELECT 1 FROM sqlite_master
                WHERE type='table' AND name='company_import_logs'
                """
            ).fetchone()
            if log_table is not None:
                connection.execute("DELETE FROM company_import_logs")
    finally:
        connection.close()


def _load_all_permits(db_path: Path) -> list[dict[str, object]]:
    planning = load_planning_permit_dataset(
        db_path,
        PROJECT_ROOT / "data" / "cloud" / "planning_construction_permits.json",
        region_key="320684",
    )
    land = load_official_permit_dataset(
        db_path,
        PROJECT_ROOT / "data" / "cloud" / "planning_land_permits.json",
        permit_type="建设用地规划许可证",
        region_key="320684",
    )
    start = load_official_permit_dataset(
        db_path,
        PROJECT_ROOT / "data" / "cloud" / "construction_start_permits.json",
        permit_type="建设工程施工许可证",
        region_key="320684",
    )
    return [*planning.items, *land.items, *start.items]


def _ten_distinct_enterprise_names(items: list[dict[str, object]]) -> list[str]:
    names: list[str] = []
    normalized_names: set[str] = set()
    invalid_names = {
        normalize_company_name(value)
        for value in ("未披露", "未知", "建设单位暂未披露", "None", "null")
    }
    for item in items:
        if str(item.get("project_type") or "") != "enterprise":
            continue
        company_name = _company_name(item)
        normalized = normalize_company_name(company_name)
        if (
            not normalized
            or normalized in invalid_names
            or normalized in normalized_names
        ):
            continue
        normalized_names.add(normalized)
        names.append(company_name)
        if len(names) == 10:
            return names
    raise AssertionError("许可证数据中不足10家可用于导入联动测试的企业")


class CompanyRegistryImportIntegrationTest(unittest.TestCase):
    def test_downloadable_template_has_all_nine_requested_headers(self):
        self.assertTrue(TEMPLATE_PATH.exists())
        workbook = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
        try:
            self.assertEqual(
                [cell.value for cell in workbook["企业工商信息导入"][1]],
                HEADERS,
            )
            self.assertIn("填写说明", workbook.sheetnames)
        finally:
            workbook.close()

    def test_import_ten_real_permit_names_recalculates_full_rule_chain(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            test_db = Path(temp_dir) / "enterprise.db"
            _copy_database(PRODUCTION_DB, test_db)
            _clear_registry_fixture(test_db)
            permit_items = _load_all_permits(test_db)
            self.assertEqual(len(permit_items), 225)
            company_names = _ten_distinct_enterprise_names(permit_items)

            result = import_company_registry_excel(
                test_db,
                _workbook_bytes(company_names),
            )
            self.assertEqual(result.total_count, 10)
            self.assertEqual(result.inserted_count, 10)
            self.assertEqual(result.updated_count, 0)
            self.assertEqual(len(list_company_registry_records(test_db)), 10)

            registry_items = enrich_items_with_company_registry(permit_items, test_db)
            coverage = summarize_registry_coverage(registry_items)
            self.assertEqual(coverage.total_project_count, 225)
            self.assertGreaterEqual(coverage.matched_project_count, 10)
            self.assertEqual(coverage.matched_company_count, 10)
            self.assertAlmostEqual(
                coverage.coverage_percentage,
                round(coverage.matched_project_count * 100 / 225, 1),
            )

            completeness_items = [
                enrich_registry_completeness(item) for item in registry_items
            ]
            strength_items = [
                enrich_company_strength(item, today=date(2026, 8, 10))
                for item in completeness_items
            ]
            finance_items = enrich_finance_opportunities(
                strength_items,
                today=date(2026, 8, 10),
            )
            selected = next(
                item
                for item in finance_items
                if normalize_company_name(_company_name(item))
                == normalize_company_name(company_names[0])
            )
            selected_raw = next(
                item
                for item in permit_items
                if normalize_company_name(_company_name(item))
                == normalize_company_name(company_names[0])
            )
            baseline = enrich_company_strength(
                enrich_registry_completeness(
                    dict(selected_raw)
                    | {
                        "registry_data_available": False,
                        "registry_disclosed_fields": [],
                    }
                ),
                today=date(2026, 8, 10),
            )
            baseline_score = score_finance_opportunity(
                baseline,
                today=date(2026, 8, 10),
            ).finance_score

            profile = build_enhanced_enterprise_profile(selected)
            report = build_marketing_report(selected, today=date(2026, 8, 10))

            self.assertEqual(selected["registry_completeness_percentage"], 100)
            self.assertEqual(selected["registry_completeness_level"], "A")
            self.assertNotEqual(selected["enterprise_strength_level"], "D")
            self.assertGreater(selected["finance_score"], baseline_score)
            self.assertTrue(profile.legal_person.startswith("测试法人"))
            self.assertEqual(profile.registered_capital, "1亿元")
            self.assertEqual(profile.company_status, "存续")
            self.assertEqual(report.company_name, profile.company_name)
            self.assertEqual(report.finance_score, selected["finance_score"])
            self.assertEqual(len(report.sections), 8)
            self.assertTrue(
                any("工商完整度=10/10" in basis for basis in report.explanation_basis)
            )


if __name__ == "__main__":
    unittest.main()
